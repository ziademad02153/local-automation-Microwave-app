"""
Excel Writer Module - UPDATED
Handles Excel file creation and data export with Pass/Fail results
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os
import config

class ExcelWriter:
    def __init__(self):
        self.workbook = None
        self.worksheet = None
    
    def create_workbook(self):
        """Create new Excel workbook"""
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Test Data"
        
        # Write headers
        self._write_headers()
    
    def _write_headers(self):
        """Write column headers with formatting"""
        headers = config.EXCEL_COLUMNS
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_num, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths
        self.worksheet.column_dimensions['A'].width = 6   # H
        self.worksheet.column_dimensions['B'].width = 6   # Min
        self.worksheet.column_dimensions['C'].width = 6   # Sec
        self.worksheet.column_dimensions['D'].width = 6   # ms
        self.worksheet.column_dimensions['E'].width = 12  # Microwave
        self.worksheet.column_dimensions['F'].width = 12  # Lamp
        self.worksheet.column_dimensions['G'].width = 12  # Door_SW
        self.worksheet.column_dimensions['H'].width = 12  # Buzzer
        self.worksheet.column_dimensions['I'].width = 12  # Grill
        self.worksheet.column_dimensions['J'].width = 12  # MW_Power%
        self.worksheet.column_dimensions['K'].width = 12  # Grill_Power%
    
    def write_data(self, data_list):
        """Write data rows to Excel"""
        if not self.worksheet:
            self.create_workbook()
        
        # Data style
        data_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row_num, data_row in enumerate(data_list, 2):  # Start from row 2
            col_num = 1
            
            # Time columns
            self.worksheet.cell(row=row_num, column=col_num, value=data_row['H']).alignment = data_alignment
            col_num += 1
            self.worksheet.cell(row=row_num, column=col_num, value=data_row['Min']).alignment = data_alignment
            col_num += 1
            self.worksheet.cell(row=row_num, column=col_num, value=data_row['Sec']).alignment = data_alignment
            col_num += 1
            self.worksheet.cell(row=row_num, column=col_num, value=data_row['ms']).alignment = data_alignment
            col_num += 1
            
            # Voltage columns (format to 3 decimals)
            for channel in ['Microwave', 'Lamp', 'Door_SW', 'Buzzer', 'Grill']:
                cell = self.worksheet.cell(row=row_num, column=col_num)
                cell.value = round(data_row[channel], 3)
                cell.alignment = data_alignment
                cell.number_format = '0.000'
                col_num += 1
            
            # Power% columns (format to 1 decimal)
            for power_col in ['MW_Power%', 'Grill_Power%']:
                cell = self.worksheet.cell(row=row_num, column=col_num)
                cell.value = round(data_row[power_col], 1)
                cell.alignment = data_alignment
                cell.number_format = '0.0'
                col_num += 1
    
    def add_summary_sheet(self, stats, test_info, pass_fail_results=None):
        """Add summary sheet with test information and Pass/Fail results"""
        summary_sheet = self.workbook.create_sheet("Summary", 0)
        
        # Title
        summary_sheet['A1'] = "Test Summary"
        summary_sheet['A1'].font = Font(bold=True, size=14)
        
        row = 3
        
        # Test information
        info_items = [
            ("Test Date:", test_info.get('date', 'N/A')),
            ("Test Mode:", test_info.get('mode', 'N/A')),
            ("Duration:", test_info.get('duration', 'N/A')),
            ("Total Samples:", test_info.get('samples', 'N/A')),
            ("", ""),
            ("MW Average Power:", f"{stats.get('mw_avg_power', 0):.1f}%"),
            ("Grill Average Power:", f"{stats.get('grill_avg_power', 0):.1f}%"),
            ("Door Opens:", stats.get('door_opens', 0)),
        ]
        
        for label, value in info_items:
            summary_sheet.cell(row=row, column=1, value=label).font = Font(bold=True)
            summary_sheet.cell(row=row, column=2, value=value)
            row += 1
        
        # Pass/Fail Results - NEW
        if pass_fail_results:
            row += 2
            summary_sheet.cell(row=row, column=1, value="Test Result:").font = Font(bold=True, size=12)
            
            overall_result = pass_fail_results.get('overall_result', 'N/A')
            result_cell = summary_sheet.cell(row=row, column=2, value=overall_result)
            result_cell.font = Font(bold=True, size=12, color="FFFFFF")
            
            if overall_result == 'PASS':
                result_cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            else:
                result_cell.fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
            
            row += 2
            
            # Details
            for detail in pass_fail_results.get('details', []):
                summary_sheet.cell(row=row, column=1, value=detail)
                row += 1
        
        # Defrost sectors if applicable
        if test_info.get('defrost_sectors'):
            row += 2
            summary_sheet.cell(row=row, column=1, value="Defrost Sectors:").font = Font(bold=True)
            row += 1
            
            summary_sheet.cell(row=row, column=1, value="Sector").font = Font(bold=True)
            summary_sheet.cell(row=row, column=2, value="Expected Power").font = Font(bold=True)
            summary_sheet.cell(row=row, column=3, value="Time Range").font = Font(bold=True)
            row += 1
            
            for sector in test_info['defrost_sectors']:
                summary_sheet.cell(row=row, column=1, value=sector['name'])
                summary_sheet.cell(row=row, column=2, value=f"{sector['expected_power']}%")
                
                start_min = int(sector['start_time'] // 60)
                start_sec = int(sector['start_time'] % 60)
                end_min = int(sector['end_time'] // 60)
                end_sec = int(sector['end_time'] % 60)
                
                summary_sheet.cell(row=row, column=3, value=f"{start_min}:{start_sec:02d} - {end_min}:{end_sec:02d}")
                row += 1
        
        # Adjust column widths
        summary_sheet.column_dimensions['A'].width = 25
        summary_sheet.column_dimensions['B'].width = 20
        summary_sheet.column_dimensions['C'].width = 20
    
    def save(self, filename):
        """Save Excel file with error handling"""
        try:
            if not self.workbook:
                return False, "No data to save"
            
            print(f"Saving workbook to: {filename}")
            
            # Save file
            self.workbook.save(filename)
            
            print(f"Workbook saved successfully")
            
            # Verify file was created
            if os.path.exists(filename):
                size = os.path.getsize(filename)
                print(f"File verified - Size: {size} bytes")
                return True, f"File saved: {filename}"
            else:
                print(f"ERROR: File not found after save!")
                return False, "File was not created"
        
        except PermissionError:
            error_msg = "Permission denied - File may be open in Excel"
            print(f"ERROR: {error_msg}")
            return False, error_msg
        
        except Exception as e:
            error_msg = f"Error saving file: {str(e)}"
            print(f"ERROR: {error_msg}")
            import traceback
            traceback.print_exc()
            return False, error_msg